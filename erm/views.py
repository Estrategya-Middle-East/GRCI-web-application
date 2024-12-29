from django.shortcuts import render
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from collections import namedtuple
from .forms import *
from .models import *
from .filters import RiskFilter
from django.db.models import Count,F,Q,Sum
from main_app.models import *
import pandas as pd
from django.http import HttpResponse,JsonResponse
from reportlab.pdfgen import canvas
import pdfkit
from django.template.loader import render_to_string
from django.apps import apps
from datetime import datetime
from django.views.generic import TemplateView
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ERM Dashboard View
def dashboard(request):
    components = [
        {"name": "Leadership and Organizational Foundations", "icon": "fas fa-briefcase", "link": "leadership/"},
        {"name": "Strategic Planning and Goal Alignment", "icon": "fas fa-chart-line", "link": "strategic_planning/"},
        {"name": "Risk Operations and Execution", "icon": "fas fa-bullseye", "link": "risks/"},
        {"name": "Continuous Monitoring and Optimization", "icon": "fas fa-balance-scale", "link": "continuous_monitoring/"},
        {"name": "Risk Management Dashboard", "icon": "fas fa-cogs", "link": "risk_intelligence/"},
    ]
    context = {
        'page_title': "ERM",
        'components': components,
    }
    return render(request, 'erm/erm_dashboard.html', context)

def leadership_dashboard(request):
    components = [
        {"name": "Board Oversight Dashboard", "icon": "fas fa-chalkboard-teacher", "link": "board_oversights"},
        {"name": "Operating Structures Manager", "icon": "fas fa-sitemap", "link": "operating_structures"},
        {"name": "Culture and Values Module", "icon": "fas fa-users", "link": "culture_surveys"},
        {"name": "Core Values Monitoring", "icon": "fas fa-compass", "link": "corevalues_monitorings"},
        {"name": "Talent Management for Risk", "icon": "fas fa-user-tie", "link": "talent_managements"},
    ]
    context = {
        'page_title': "Leadership",
        'components': components,
    }
    return render(request, 'leadership/leadership_dashboard.html', context)


def strategic_planning_dashboard(request):
    components = [
        {"name": "Business Context Analyzer", "icon": "fas fa-chart-area", "link": "business_contexts/"},
        {"name": "Risk Appetite Manager", "icon": "fas fa-adjust", "link": "risk_appetites/"},
        {"name": "Strategic Evaluation Tool", "icon": "fas fa-search-plus", "link": "strategic_evaluations/"},
        {"name": "Objective Setting Dashboard", "icon": "fas fa-bullseye", "link": "objectives/"},
    ]
    context = {
        'page_title': "Strategic Planning Dashboard",
        'components': components,
    }
    return render(request, 'strategic_planning/dashboard.html', context)


def continuous_monitoring_dashboard(request):
    components = [
        {"name": "Change Impact Assessor", "icon": "fas fa-exchange-alt", "link": "change_assessments"},
        {"name": "Performance Review Module", "icon": "fas fa-chart-bar", "link": "performance_reviews"},
        {"name": "Continuous Improvement Tool", "icon": "fas fa-tools", "link": "improvement_actions"},
    ]
    context = {
        'page_title': "Continuous Monitoring and Optimization",
        'components': components,
    }
    return render(request, 'continuous_monitoring/dashboard.html', context)


def risk_intelligence_dashboard(request):
    inherent_risk_counts = (
        RiskDefine.objects.values("category")
        .annotate(count=Count("category"))
        .order_by("category")
    )
    
     # Calculate total inherent risks
    total_inherent_risks = sum(item["count"] for item in inherent_risk_counts)
    
    # Prepare the data for ECharts
    chart_data_inherent = [{"name": item["category"], "value": item["count"]} for item in inherent_risk_counts]
    
    residual_risks = (
        RiskResidualAss.objects.filter(risk_score__gt=0)
        .values("risk__define_step__category")
        .annotate(count=Count("risk__define_step__category"))
    )
   
    # Calculate total residual risks
    total_residual_risks = sum(item["count"] for item in residual_risks)
    
    chart_data_residual = [
        {"name": item["risk__define_step__category"], "value": item["count"]} for item in residual_risks
    ]
    
    # Aggregate data: Count risks by likelihood and impact
    heatmap_data = (
        RiskAss.objects
        .values('likelihood_rating', 'impact_rating', 'risk_score')  # Group by likelihood and impact
        .annotate(count=Count('id'))  # Count the number of risks
    )

    # Format data for ECharts: [impact-1, 5-likelihood, count]
    formatted_data = [
        [item['impact_rating'] - 1, 5 - item['likelihood_rating'], item['count'], item['risk_score']]
        for item in heatmap_data
    ]
    
    residual_heatmap_data = (
        RiskResidualAss.objects
        .values('likelihood_rating', 'impact_rating', 'risk_score')  # Group by likelihood and impact
        .annotate(count=Count('id'))  # Count the number of risks
    )

    # Format data for ECharts: [impact-1, 5-likelihood, count]
    residual_formatted_data = [
        [item['impact_rating'] - 1, 5 - item['likelihood_rating'], item['count'], item['risk_score']]
        for item in residual_heatmap_data
    ]
    
    
    # risk score ranges
    high_risk_range = Q(risk_score__gte=20, risk_score__lte=25)
    risk_tolerance_range = Q(risk_score__gte=10, risk_score__lt=20)
    risk_appetite_range = Q(risk_score__lt=10)

    # Calculate counts for each risk range
    high_risk_count = RiskAss.objects.filter(high_risk_range).count()
    risk_tolerance_count = RiskAss.objects.filter(risk_tolerance_range).count()
    risk_appetite_count = RiskAss.objects.filter(risk_appetite_range).count()

    # Total risks
    inherent_gauge_total_risks = high_risk_count + risk_tolerance_count + risk_appetite_count

    # Calculate percentages
    inherent_gauge_chart_data = [
        {"value": round((high_risk_count / inherent_gauge_total_risks) * 100, 2), "name": "High-Risk Range"},
        {"value": round((risk_tolerance_count / inherent_gauge_total_risks) * 100, 2), "name": "Risk Tolerance"},
        {"value": round((risk_appetite_count / inherent_gauge_total_risks) * 100, 2), "name": "Risk Appetite"},
    ]

    
    # Calculate counts for each risk range
    high_residual_risk_count = RiskResidualAss.objects.filter(high_risk_range).count()
    residual_risk_tolerance_count = RiskResidualAss.objects.filter(risk_tolerance_range).count()
    residual_risk_appetite_count = RiskResidualAss.objects.filter(risk_appetite_range).count()

    # Total risks
    residual_gauge_risks = high_residual_risk_count + residual_risk_tolerance_count + residual_risk_appetite_count

    # Calculate percentages
    residual_gauge_chart_data = [
        {"value": round((high_residual_risk_count / residual_gauge_risks) * 100, 2), "name": "High-Risk Range"},
        {"value": round((residual_risk_tolerance_count / residual_gauge_risks) * 100, 2), "name": "Risk Tolerance"},
        {"value": round((residual_risk_appetite_count / residual_gauge_risks) * 100, 2), "name": "Risk Appetite"},
    ]
    
    department_totals = RiskDefine.objects.values('department__name').annotate(
    total_impact=Sum('impact'),
    total_likelihood=Sum('likelihood'),
    total_risk_score=Sum('risk_score')
    )

    # Format data for the chart
    department_data = {
        'departments': [dept['department__name'] or 'Unknown' for dept in department_totals],
        'impact_totals': [dept['total_impact'] or 0 for dept in department_totals],
        'likelihood_totals': [dept['total_likelihood'] or 0 for dept in department_totals],
        'risk_scores': [dept['total_risk_score'] or 0 for dept in department_totals],
    }
    
    # Organizational structure tree logic
    levels = ['N1', 'N2', 'N3', 'N4']
    org_chart_data = []

    # Add CEO as the root node
    ceo_node = {
        "name": "CEO",
        "value": 0,
        "children": []
    }

    # Build hierarchy for each level
    current_level_nodes = [ceo_node]  # Start from CEO

    for level in levels:
        departments = Department.objects.filter(org_chart_level=level).annotate(staff_count=Count('staff'))
        next_level_nodes = []

        for department in departments:
            department_node = {
                "name": department.name,
                "value": department.staff_count,
                "tooltip": {
                    "formatter": f"{department.name}: {department.staff_count} staff"
                },
                "children": []
            }
            # Add this node to the parent's children
            for parent in current_level_nodes:
                if parent["name"] == "CEO" or parent.get("children"):
                    parent["children"].append(department_node)
            next_level_nodes.append(department_node)

        # Update the current level nodes for the next iteration
        current_level_nodes = next_level_nodes

    # Add CEO node to the data
    org_chart_data.append(ceo_node)
    
    context = {
        'page_title': "Risk Management Dashboard",
        "chart_data_inherent": chart_data_inherent,
        'total_inherent_risks': total_inherent_risks,
        'heatmap_data': formatted_data,
        'inherent_gauge_chart_data': inherent_gauge_chart_data,
        'chart_data_residual': chart_data_residual,
        'total_residual_risks':total_residual_risks,
        'residual_heatmap_data': residual_formatted_data,
        'residual_gauge_chart_data': residual_gauge_chart_data,
        'department_chart_data': department_data,
        'org_chart_data': org_chart_data,  # Pass organizational structure data
    }
    return render(request, 'risk_intelligence/dashboard.html', context)

""" 
def export_all_to_excel(request):
    # Create an Excel writer
    with pd.ExcelWriter('all_data.xlsx', engine='openpyxl') as writer:
        # Get all models in the app
        app_models = apps.get_models()
        
        for model in app_models:
            # Fetch all records of the model
            queryset = model.objects.all()
            
            # If there are no records, skip this model
            if not queryset.exists():
                continue
            
            # Get the model name
            model_name = model.__name__
            
            # Serialize the data into a list of dictionaries
            data = []
            for obj in queryset:
                record = {}
                for field in model._meta.fields:
                    field_name = field.name
                    value = getattr(obj, field_name)
                    
                    # Handle foreign key relationships
                    if field.is_relation and value:
                        related_field = field.related_model._meta.verbose_name
                        value = str(value)  # Convert related object to string (or name field if needed)
                    
                    # Handle timezone-aware datetimes
                    if isinstance(value, datetime) and value.tzinfo is not None:
                        value = value.astimezone(None).strftime('%Y-%m-%d %H:%M:%S')
                    
                    record[field.verbose_name] = value
                data.append(record)
            
            # Convert to a DataFrame
            df = pd.DataFrame(data)
            
            # Write to Excel (each model gets a separate sheet)
            df.to_excel(writer, index=False, sheet_name=model_name[:30])  # Sheet names max 31 chars
            
    # Prepare the file for download
    with open('all_data.xlsx', 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="all_data.xlsx"'
        return response
 """
 
def export_risks_to_excel(request):
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risks"

    # Define headers
    headers = [
        "Risk Name", "Description",
        "Department", "Objective", "Identified By", "Identification Date",
        "Category", "Subcategory", "Source", "Risk Cause", "Risk Event", "Risk Score", 
        "Likelihood Rating", "Impact Rating", "Risk Score",  "Risk Heatmap Position", "Approval Status"
    ]

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment_center = Alignment(horizontal="left", vertical="center")
    alt_row_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    # Apply header styles
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border_style

    # Query data from Risk, RiskDefine, and RiskAss models
    risks = Risk.objects.all()
    row = 2
    total_risk_score = 0  # Initialize total risk score

    for index, risk in enumerate(risks):
        define_step = getattr(risk, 'define_step', None)
        assessment_step = getattr(risk, 'assessment_step', None)

        # Alternate row background for better readability
        fill_color = alt_row_fill if row % 2 == 0 else None

        # Safely extract foreign key fields and handle None values
        department = str(getattr(define_step, 'department', '')) if define_step and define_step.department else ''
        identified_by = str(getattr(define_step, 'identified_by', '')) if define_step and define_step.identified_by else ''
        objective = str(getattr(define_step, 'objective', '')) if define_step and define_step.objective else ''

        # Data Rows
        cells = [
            risk.name, 
            risk.description, 
            department, 
            objective,
            identified_by, 
            getattr(define_step, 'identification_date', '').strftime('%Y-%m-%d') if define_step and define_step.identification_date else '',
            getattr(define_step, 'category', ''),
            getattr(define_step, 'subcategory', ''),
            getattr(define_step, 'source', ''),
            getattr(define_step, 'risk_cause', ''),
            getattr(define_step, 'risk_event', ''),
            getattr(define_step, 'risk_impact', ''),
            getattr(assessment_step, 'likelihood_rating', ''),
            getattr(assessment_step, 'impact_rating', ''),
            getattr(assessment_step, 'risk_score', ''),
            getattr(assessment_step, 'risk_heatmap_position', ''),
            getattr(define_step, 'approval_status', '')
        ]

        # Add risk score to total
        risk_score = getattr(define_step, 'risk_score', 0)
        total_risk_score += risk_score if isinstance(risk_score, (int, float)) else 0

        for col_num, value in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=col_num, value=value)
            cell.border = border_style
            if fill_color:
                cell.fill = fill_color
            cell.font = Font(size=10)
        row += 1

    # Adjust column widths dynamically
    for col_num, col in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    # Add a summary row at the end
    summary_row = row + 1
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=len(headers))
    summary_cell = ws.cell(row=summary_row, column=1, value=f"Total Risks: {risks.count()} | Total Risk Score: {total_risk_score}")
    summary_cell.font = Font(bold=True, size=11, color="FFFFFF")
    summary_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_cell.alignment = alignment_center

    # Prepare the HTTP response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="risks_report.xlsx"'
    wb.save(response)
    return response


class RiskWorkflowView(TemplateView):
    template_name = 'erm/workflow.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        risk_id = kwargs['id']
        risk = get_object_or_404(Risk, id=risk_id)

        # Ensure we always pass both forms, even if the step is approved or completed
        # new
        if risk.workflow_status == 'new':
            risk.workflow_status = 'define'
            risk.save()
            messages.info(self.request, "Workflow status updated to 'define'.")
        # Define step form
        define_step = RiskDefine.objects.filter(risk=risk).first()
        if not define_step:
            define_step = RiskDefine(risk=risk)
            define_step.save()
        context['form_define'] = RiskDefineForm(instance=define_step)

        # Assessment step form
        assessment_step = RiskAss.objects.filter(risk=risk).first()
        if not assessment_step:
            assessment_step = RiskAss(risk=risk)
            assessment_step.save()
        context['form_assessment'] = RiskAssForm(instance=assessment_step)
        
        # Prioritization step form
        prioritization_step = RiskPrioritization.objects.filter(risk=risk).first()
        if not prioritization_step:
            prioritization_step = RiskPrioritization(risk=risk)
            prioritization_step.save()
        context['form_prioritization'] = RiskPrioritizationForm(instance=prioritization_step)
        
        # Response step form
        response_step = RiskResponse.objects.filter(risk=risk).first()
        if not response_step:
            response_step = RiskResponse(risk=risk)
            response_step.save()
        context['form_response'] = RiskResponseForm(instance=response_step)
        
        # Residual step form
        residual_step = RiskResidualAss.objects.filter(risk=risk).first()
        if not residual_step:
            residual_step = RiskResidualAss(risk=risk)
            residual_step.save()
        context['form_residual'] = RiskResidualAssForm(instance=residual_step)
        
        # Close step form
        close_step = RiskClose.objects.filter(risk=risk).first()
        if not close_step:
            close_step = RiskClose(risk=risk)
            close_step.save()
        context['form_close'] = RiskCloseForm(instance=close_step)
            
        # Set the current step
        if risk.workflow_status == 'define':
            context['current_step'] = 'define'
        elif risk.workflow_status == 'assessment':
            context['current_step'] = 'assessment'
        elif risk.workflow_status == 'prioritization':
            context['current_step'] = 'prioritization'
        elif risk.workflow_status == 'response':
            context['current_step'] = 'response'
        elif risk.workflow_status == 'residual':
            context['current_step'] = 'residual'
        elif risk.workflow_status == 'close':
            context['current_step'] = 'close'
        
        context['risk'] = risk
        return context


    def post(self, request, *args, **kwargs):
        risk_id = kwargs['id']
        risk = get_object_or_404(Risk, id=risk_id)

        # Handle different workflow steps
        if risk.workflow_status == 'define':
            define_step = RiskDefine.objects.filter(risk=risk).first()

            form = RiskDefineForm(request.POST, instance=define_step)
            if form.is_valid():
                form.save()

                if 'save' in request.POST:
                    # Save action: update the current form data
                    messages.success(request, "Risk saved successfully!")

                elif 'submit_to_manager' in request.POST:
                    # Submit to Manager action
                    risk.workflow_status = 'define'
                    risk.define_approval_status = 'under_review'
                    define_step.approval_status = 'under_review'
                    risk.save()
                    define_step.save()
                    messages.success(request, "Risk submitted to manager.")

                elif 'approve' in request.POST:
                    # Approve action: move to next step (assessment)
                    risk.workflow_status = 'assessment'
                    risk.define_approval_status = 'approved'
                    define_step.approval_status = 'approved'
                    risk.save()
                    define_step.save()
                    messages.success(request, "Risk approved and moved to prioritization.")

                elif 'reject' in request.POST:
                    # Reject action: set status to 'under_review' and revert the buttons
                    risk.workflow_status = 'define'
                    risk.define_approval_status = 'rejected'                    
                    define_step.approval_status = 'rejected'
                    risk.save()
                    define_step.save()
                    messages.info(request, "Risk rejected.")

        elif risk.workflow_status == 'assessment':
            assessment_step = RiskAss.objects.filter(risk=risk).first()

            form = RiskAssForm(request.POST, instance=assessment_step)
            if form.is_valid():
                form.save()

                if 'save' in request.POST:
                    # Save action for assessment
                    messages.success(request, "Assessment data saved.")

                elif 'submit_to_manager' in request.POST:
                    # Submit to Manager action for assessment
                    risk.workflow_status = 'assessment'
                    risk.assessment_approval_status = 'under_review'
                    assessment_step.approval_status = 'under_review'
                    risk.save()
                    assessment_step.save()
                    messages.success(request, "Assessment submitted to manager.")

                elif 'approve' in request.POST:
                    # Approve action for assessment: move to the next step (e.g., control or closure)
                    risk.workflow_status = 'prioritization'  # Assuming next step is 'control'
                    risk.assessment_approval_status = 'approved'
                    assessment_step.approval_status = 'approved'
                    risk.save()
                    assessment_step.save()
                    messages.success(request, "Assessment approved and moved to control.")

                elif 'reject' in request.POST:
                    # Reject action for assessment: revert to under review
                    risk.workflow_status = 'assessment'
                    risk.assessment_approval_status = 'rejected'
                    assessment_step.approval_status = 'rejected'
                    risk.save()
                    assessment_step.save()
                    messages.info(request, "Assessment rejected.")

        elif risk.workflow_status == 'prioritization':
            prioritization_step = RiskPrioritization.objects.filter(risk=risk).first()

            form = RiskPrioritizationForm(request.POST, instance=prioritization_step)
            if form.is_valid():
                form.save()

                if 'save' in request.POST:
                    # Save action for prioritization
                    messages.success(request, "Prioritization data saved.")

                elif 'submit_to_manager' in request.POST:
                    # Submit to Manager action for prioritization
                    risk.workflow_status = 'prioritization'
                    risk.prioritization_approval_status = 'under_review'
                    prioritization_step.approval_status = 'under_review'
                    risk.save()
                    prioritization_step.save()
                    messages.success(request, "Prioritization submitted to manager.")

                elif 'approve' in request.POST:
                    # Approve action for prioritization: move to the next step 
                    risk.workflow_status = 'response'  
                    risk.prioritization_approval_status = 'approved'
                    prioritization_step.approval_status = 'approved'
                    risk.save()
                    prioritization_step.save()
                    messages.success(request, "Prioritization approved.")

                elif 'reject' in request.POST:
                    # Reject action for prioritization: revert to under review
                    risk.workflow_status = 'prioritization'
                    risk.prioritization_approval_status = 'rejected'
                    prioritization_step.approval_status = 'rejected'
                    risk.save()
                    prioritization_step.save()
                    messages.info(request, "Prioritization rejected.")

        elif risk.workflow_status == 'response':
            response_step = RiskResponse.objects.filter(risk=risk).first()

            form = RiskResponseForm(request.POST, instance=response_step)
            if form.is_valid():
                form.save()

                if 'save' in request.POST:
                    # Save action for response
                    messages.success(request, "Response data saved.")

                elif 'submit_to_manager' in request.POST:
                    # Submit to Manager action for response
                    risk.workflow_status = 'response'
                    risk.response_approval_status = 'under_review'
                    response_step.approval_status = 'under_review'
                    risk.save()
                    response_step.save()
                    messages.success(request, "Response submitted to manager.")

                elif 'approve' in request.POST:
                    # Approve action for response: move to the next step 
                    risk.workflow_status = 'residual'  
                    risk.response_approval_status = 'approved'
                    response_step.approval_status = 'approved'
                    risk.save()
                    response_step.save()
                    messages.success(request, "Response approved.")

                elif 'reject' in request.POST:
                    # Reject action for response: revert to under review
                    risk.workflow_status = 'response'
                    risk.response_approval_status = 'rejected'
                    response_step.approval_status = 'rejected'
                    risk.save()
                    response_step.save()
                    messages.info(request, "Response rejected.")
        
        elif risk.workflow_status == 'residual':
            residual_step = RiskResidualAss.objects.filter(risk=risk).first()

            form = RiskResidualAssForm(request.POST, instance=residual_step)
            if form.is_valid():
                form.save()

                if 'save' in request.POST:
                    # Save action for residual
                    messages.success(request, "Residual data saved.")

                elif 'submit_to_manager' in request.POST:
                    # Submit to Manager action for residual
                    risk.workflow_status = 'residual'
                    risk.residual_approval_status = 'under_review'
                    residual_step.approval_status = 'under_review'
                    risk.save()
                    residual_step.save()
                    messages.success(request, "Residual submitted to manager.")

                elif 'approve' in request.POST:
                    # Approve action for residual: move to the next step 
                    risk.workflow_status = 'close'  
                    risk.residual_approval_status = 'approved'
                    residual_step.approval_status = 'approved'
                    risk.save()
                    residual_step.save()
                    messages.success(request, "Residual approved.")

                elif 'reject' in request.POST:
                    # Reject action for residual: revert to under review
                    risk.workflow_status = 'residual'
                    risk.residual_approval_status = 'rejected'
                    residual_step.approval_status = 'rejected'
                    risk.save()
                    residual_step.save()
                    messages.info(request, "Residual rejected.")
            
        elif risk.workflow_status == 'close':
            close_step = RiskClose.objects.filter(risk=risk).first()

            form = RiskCloseForm(request.POST, instance=close_step)
            if form.is_valid():
                form.save()

                if 'save' in request.POST:
                    # Save action for close
                    messages.success(request, "Close data saved.")

                elif 'submit_to_manager' in request.POST:
                    # Submit to Manager action for close
                    risk.workflow_status = 'close'
                    risk.close_approval_status = 'under_review'
                    close_step.approval_status = 'under_review'
                    risk.save()
                    close_step.save()
                    messages.success(request, "Close submitted to manager.")

                elif 'approve' in request.POST:
                    # Approve action for close: move to the next step 
                    risk.workflow_status = 'completed'  
                    risk.close_approval_status = 'approved'
                    close_step.approval_status = 'approved'
                    risk.save()
                    close_step.save()
                    messages.success(request, "close approved.")

                elif 'reject' in request.POST:
                    # Reject action for close: revert to under review
                    risk.workflow_status = 'close'
                    risk.close_approval_status = 'rejected'
                    close_step.approval_status = 'rejected'
                    risk.save()
                    close_step.save()
                    messages.info(request, "Close rejected.")

        return redirect('workflow_view', id=risk.id)
    
   
def list_risks(request):
    # Get all records by default
    risks = Risk.objects.all()
    form = RiskForm()

    # Search filter
    # Apply the filter
    risk_filter = RiskFilter(request.GET, queryset=risks)
    
    
    # Default sorting by ID
    sort_by = request.GET.get('sort_by', 'id')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        risks = risks.order_by(sort_by)
    else:
        risks = risks.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(risks, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_risks = paginator.page(page)
    except PageNotAnInteger:
        paginated_risks = paginator.page(1)
    except EmptyPage:
        paginated_risks = paginator.page(paginator.num_pages)

    risks = risk_filter.qs
    
    inherent_risk_counts = (
        RiskDefine.objects.values("category")
        .annotate(count=Count("category"))
        .order_by("category")
    )
    
     # Calculate total inherent risks
    total_inherent_risks = sum(item["count"] for item in inherent_risk_counts)
    
    # Prepare the data for ECharts
    chart_data_inherent = [{"name": item["category"], "value": item["count"]} for item in inherent_risk_counts]
    
    residual_risks = (
        RiskResidualAss.objects.filter(risk_score__gt=0)
        .values("risk__define_step__category")
        .annotate(count=Count("risk__define_step__category"))
    )
   
    # Calculate total residual risks
    total_residual_risks = sum(item["count"] for item in residual_risks)
    
    chart_data_residual = [
        {"name": item["risk__define_step__category"], "value": item["count"]} for item in residual_risks
    ]
    
    # Aggregate data: Count risks by likelihood and impact
    heatmap_data = (
        RiskAss.objects
        .values('likelihood_rating', 'impact_rating', 'risk_score')  # Group by likelihood and impact
        .annotate(count=Count('id'))  # Count the number of risks
    )

    # Format data for ECharts: [impact-1, 5-likelihood, count]
    formatted_data = [
        [item['impact_rating'] - 1, 5 - item['likelihood_rating'], item['count'], item['risk_score']]
        for item in heatmap_data
    ]
    
    residual_heatmap_data = (
        RiskResidualAss.objects
        .values('likelihood_rating', 'impact_rating', 'risk_score')  # Group by likelihood and impact
        .annotate(count=Count('id'))  # Count the number of risks
    )

    # Format data for ECharts: [impact-1, 5-likelihood, count]
    residual_formatted_data = [
        [item['impact_rating'] - 1, 5 - item['likelihood_rating'], item['count'], item['risk_score']]
        for item in residual_heatmap_data
    ]
    
    
    # risk score ranges
    high_risk_range = Q(risk_score__gte=20, risk_score__lte=25)
    risk_tolerance_range = Q(risk_score__gte=10, risk_score__lt=20)
    risk_appetite_range = Q(risk_score__lt=10)

    # Calculate counts for each risk range
    high_risk_count = RiskAss.objects.filter(high_risk_range).count()
    risk_tolerance_count = RiskAss.objects.filter(risk_tolerance_range).count()
    risk_appetite_count = RiskAss.objects.filter(risk_appetite_range).count()

    # Total risks
    inherent_gauge_total_risks = high_risk_count + risk_tolerance_count + risk_appetite_count

    # Calculate percentages
    inherent_gauge_chart_data = [
        {"value": round((high_risk_count / inherent_gauge_total_risks) * 100, 2), "name": "High-Risk Range"},
        {"value": round((risk_tolerance_count / inherent_gauge_total_risks) * 100, 2), "name": "Risk Tolerance"},
        {"value": round((risk_appetite_count / inherent_gauge_total_risks) * 100, 2), "name": "Risk Appetite"},
    ]

    
    # Calculate counts for each risk range
    high_residual_risk_count = RiskResidualAss.objects.filter(high_risk_range).count()
    residual_risk_tolerance_count = RiskResidualAss.objects.filter(risk_tolerance_range).count()
    residual_risk_appetite_count = RiskResidualAss.objects.filter(risk_appetite_range).count()

    # Total risks
    residual_gauge_risks = high_residual_risk_count + residual_risk_tolerance_count + residual_risk_appetite_count

    # Calculate percentages
    residual_gauge_chart_data = [
        {"value": round((high_residual_risk_count / residual_gauge_risks) * 100, 2), "name": "High-Risk Range"},
        {"value": round((residual_risk_tolerance_count / residual_gauge_risks) * 100, 2), "name": "Risk Tolerance"},
        {"value": round((residual_risk_appetite_count / residual_gauge_risks) * 100, 2), "name": "Risk Appetite"},
    ]
    
    department_totals = RiskDefine.objects.values('department__name').annotate(
    total_impact=Sum('impact'),
    total_likelihood=Sum('likelihood'),
    total_risk_score=Sum('risk_score')
    )

    # Format data for the chart
    department_data = {
        'departments': [dept['department__name'] or 'Unknown' for dept in department_totals],
        'impact_totals': [dept['total_impact'] or 0 for dept in department_totals],
        'likelihood_totals': [dept['total_likelihood'] or 0 for dept in department_totals],
        'risk_scores': [dept['total_risk_score'] or 0 for dept in department_totals],
    }
    
    # Pass necessary context to the template
    context = {
        'page_title': "Risks",
        'risks':risks, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_risks': paginated_risks,
        'rows_per_page': rows_per_page,
        'form': form,
        'risk_filter': risk_filter,
        "chart_data_inherent": chart_data_inherent,
        'total_inherent_risks': total_inherent_risks,
        'heatmap_data': formatted_data,
        'inherent_gauge_chart_data': inherent_gauge_chart_data,
        'chart_data_residual': chart_data_residual,
        'total_residual_risks':total_residual_risks,
        'residual_heatmap_data': residual_formatted_data,
        'residual_gauge_chart_data': residual_gauge_chart_data,
        'department_chart_data': department_data,
    }
    return render(request, 'erm/list_risks.html', context)


# Add Risk
def add_risk(request):
    form = RiskForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            risk = form.save()
            # Set the workflow status to 'new' after saving
            risk.workflow_status = 'new'
            risk.save()
            messages.success(request, "Risk added successfully!")
            return redirect('list_risks')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Risk. Please correct the errors.")
    return redirect('list_risks')

# Edit Risk
def edit_risk(request, id):
    risk = get_object_or_404(Risk, id=id)
    if request.method == 'POST':
        form = RiskForm(request.POST, instance=risk)
        if form.is_valid():
            form.save()
            messages.success(request, "Risk updated successfully!")
            return redirect('list_risks')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Risk. Please correct the errors.")
    else:
        form = RiskForm(instance=risk)
    return render(request, 'erm/list_risks.html', {'form': form})



# Delete Risk
def delete_risk(request, id):
    risk = get_object_or_404(Risk, id=id) 
    risk.delete()
    messages.success(request, "Risk deleted successfully!")
    return redirect('list_risks')


#############################
######## LEADERSHIP #########
#############################


def list_board_oversight(request):
    # Get all records by default
    board_oversights = Oversight.objects.all()
    form = OversightForm()

    
    # Default sorting by OversightID
    sort_by = request.GET.get('sort_by', 'OversightID')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        board_oversights = board_oversights.order_by(sort_by)
    else:
        board_oversights = board_oversights.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(board_oversights, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_board_oversights = paginator.page(page)
    except PageNotAnInteger:
        paginated_board_oversights = paginator.page(1)
    except EmptyPage:
        paginated_board_oversights = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Board Oversight",
        'board_oversights':board_oversights, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_board_oversights': paginated_board_oversights,
        'rows_per_page': rows_per_page,
        'form': form,
    }
    return render(request, 'leadership/board_oversight.html', context)


# Add Oversigh 

def add_board_oversight(request):
    form = OversightForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Oversight added successfully!")
            return redirect('list_board_oversight')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Oversight. Please correct the errors.")
    return redirect('list_board_oversight')

# Edit Board Oversight Dashboard
def edit_board_oversight(request, OversightID):
    board_oversight = get_object_or_404(Oversight, OversightID=OversightID)
    if request.method == 'POST':
        form = OversightForm(request.POST, instance=board_oversight)
        if form.is_valid():
            form.save()
            messages.success(request, "Board Oversight updated successfully!")
            return redirect('list_board_oversight')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Board Oversight. Please correct the errors.")
    else:
        form = OversightForm(instance=board_oversight)
    return render(request, 'leadership/board_oversight.html', {'form': form})



# Delete board_oversight
def delete_board_oversight(request, OversightID):
    board_oversight = get_object_or_404(Oversight, OversightID=OversightID) 
    board_oversight.delete()
    messages.success(request, "Board Oversight deleted successfully!")
    return redirect('list_board_oversight')

##############################
# Operating Structures Manager
##############################

def list_operating_structure(request):
    # Get all records by default
    operating_structures = OperatingStructure.objects.all()
    form = OperatingStructureForm()

    
    # Default sorting by StructureID
    sort_by = request.GET.get('sort_by', 'StructureID')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        operating_structures = operating_structures.order_by(sort_by)
    else:
        operating_structures = operating_structures.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(operating_structures, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_operating_structures = paginator.page(page)
    except PageNotAnInteger:
        paginated_operating_structures = paginator.page(1)
    except EmptyPage:
        paginated_operating_structures = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Operating Structure",
        'operating_structures':operating_structures, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_operating_structures': paginated_operating_structures,
        'rows_per_page': rows_per_page,
        'form': form,
        'Status_choices': OperatingStructure._meta.get_field('Status').choices,
    }
    return render(request, 'leadership/operating_structure.html', context)


# Add Operating Structure

def add_operating_structure(request):
    form = OperatingStructureForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Structure added successfully!")
            return redirect('list_operating_structure')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Structure. Please correct the errors.")
    return redirect('list_operating_structure')

# Edit Structure
def edit_operating_structure(request, StructureID):
    operating_structure = get_object_or_404(OperatingStructure, StructureID=StructureID)
    if request.method == 'POST':
        form = OperatingStructureForm(request.POST, instance=operating_structure)
        if form.is_valid():
            form.save()
            messages.success(request, "Structure updated successfully!")
            return redirect('list_operating_structure')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Structure. Please correct the errors.")
    else:
        form = OperatingStructureForm(instance=operating_structure)
    return render(request, 'leadership/operating_structure.html', {'form': form})



# Delete operating_structure
def delete_operating_structure(request, StructureID):
    operating_structure = get_object_or_404(OperatingStructure, StructureID=StructureID) 
    operating_structure.delete()
    messages.success(request, "Structure deleted successfully!")
    return redirect('list_operating_structure')


###############
#Culture Survey
###############

def list_culture_survey(request):
    # Get all records by default
    culture_surveys = CultureSurvey.objects.all()
    form = CultureSurveyForm()

    
    # Default sorting by SurveyID
    sort_by = request.GET.get('sort_by', 'SurveyID')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        culture_surveys = culture_surveys.order_by(sort_by)
    else:
        culture_surveys = culture_surveys.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(culture_surveys, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_culture_surveys = paginator.page(page)
    except PageNotAnInteger:
        paginated_culture_surveys = paginator.page(1)
    except EmptyPage:
        paginated_culture_surveys = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Culture and Values Survey",
        'culture_surveys':culture_surveys, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_culture_surveys': paginated_culture_surveys,
        'rows_per_page': rows_per_page,
        'form': form,
        'Respondents': Staff.objects.all(),
    }
    return render(request, 'leadership/culture_survey.html', context)


# Add Oversigh 

def add_culture_survey(request):
    form = CultureSurveyForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Survey added successfully!")
            return redirect('list_culture_survey')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Survey. Please correct the errors.")
    return redirect('list_culture_survey')

# Edit Board Survey Dashboard
def edit_culture_survey(request, SurveyID):
    culture_survey = get_object_or_404(CultureSurvey, SurveyID=SurveyID)
    if request.method == 'POST':
        form = CultureSurveyForm(request.POST, instance=culture_survey)
        if form.is_valid():
            form.save()
            messages.success(request, "Survey updated successfully!")
            return redirect('list_culture_survey')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Survey. Please correct the errors.")
    else:
        form = CultureSurveyForm(instance=culture_survey)
    return render(request, 'leadership/culture_survey.html', {'form': form})


# Delete culture_survey
def delete_culture_survey(request, SurveyID):
    culture_survey = get_object_or_404(CultureSurvey, SurveyID=SurveyID) 
    culture_survey.delete()
    messages.success(request, "Survey deleted successfully!")
    return redirect('list_culture_survey')


##############################
#Core Values Monitoring
##############################

def list_corevalues_monitoring(request):
    # Get all records by default
    corevalues_monitorings = CoreValuesMonitoring.objects.all()
    form = CoreValuesMonitoringForm()

    
    # Default sorting by MonitoringID
    sort_by = request.GET.get('sort_by', 'MonitoringID')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        corevalues_monitorings = corevalues_monitorings.order_by(sort_by)
    else:
        corevalues_monitorings = corevalues_monitorings.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(corevalues_monitorings, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_corevalues_monitorings = paginator.page(page)
    except PageNotAnInteger:
        paginated_corevalues_monitorings = paginator.page(1)
    except EmptyPage:
        paginated_corevalues_monitorings = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Core Values Monitoring",
        'corevalues_monitorings':corevalues_monitorings, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_corevalues_monitorings': paginated_corevalues_monitorings,
        'rows_per_page': rows_per_page,
        'form': form,
        'ReportedBys' : Staff.objects.all(),       
        'Status_choices': CoreValuesMonitoring._meta.get_field('ComplianceStatus').choices,
    }
    return render(request, 'leadership/corevalues_monitoring.html', context)


# Add Oversigh 

def add_corevalues_monitoring(request):
    form = CoreValuesMonitoringForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Core Values Monitoring added successfully!")
            return redirect('list_corevalues_monitoring')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Core Values Monitoring. Please correct the errors.")
    return redirect('list_corevalues_monitoring')

# Edit Board Monitoring Dashboard
def edit_corevalues_monitoring(request, MonitoringID):
    corevalues_monitoring = get_object_or_404(CoreValuesMonitoring, MonitoringID=MonitoringID)
    if request.method == 'POST':
        form = CoreValuesMonitoringForm(request.POST, instance=corevalues_monitoring)
        if form.is_valid():
            form.save()
            messages.success(request, "Core Values Monitoring updated successfully!")
            return redirect('list_corevalues_monitoring')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Core Values Monitoring. Please correct the errors.")
    else:
        form = CoreValuesMonitoringForm(instance=corevalues_monitoring)
    return render(request, 'leadership/corevalues_monitoring.html', {'form': form})


# Delete corevalues_monitoring
def delete_corevalues_monitoring(request, MonitoringID):
    corevalues_monitoring = get_object_or_404(CoreValuesMonitoring, MonitoringID=MonitoringID) 
    corevalues_monitoring.delete()
    messages.success(request, "Core Values Monitoring deleted successfully!")
    return redirect('list_corevalues_monitoring')

##############################
# Talent Management for Risk
##############################

def list_talent_management(request):
    # Get all records by default
    talent_managements = RiskTalent.objects.all()
    form = RiskTalentForm()

    
    # Default sorting by TalentID
    sort_by = request.GET.get('sort_by', 'TalentID')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        talent_managements = talent_managements.order_by(sort_by)
    else:
        talent_managements = talent_managements.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(talent_managements, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_talent_managements = paginator.page(page)
    except PageNotAnInteger:
        paginated_talent_managements = paginator.page(1)
    except EmptyPage:
        paginated_talent_managements = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Risk Talent",
        'talent_managements':talent_managements, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_talent_managements': paginated_talent_managements,
        'rows_per_page': rows_per_page,
        'form': form,
    }
    return render(request, 'leadership/talent_management.html', context)


# Add Oversigh 

def add_talent_management(request):
    form = RiskTalentForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Talent added successfully!")
            return redirect('list_talent_management')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Risk Talent. Please correct the errors.")
    return redirect('list_talent_management')

# Edit Board RiskTalent Dashboard
def edit_talent_management(request, TalentID):
    talent_management = get_object_or_404(RiskTalent, TalentID=TalentID)
    if request.method == 'POST':
        form = RiskTalentForm(request.POST, instance=talent_management)
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Talent updated successfully!")
            return redirect('list_talent_management')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Risk Talent. Please correct the errors.")
    else:
        form = RiskTalentForm(instance=talent_management)
    return render(request, 'leadership/talent_management.html', {'form': form})


# Delete talent_management
def delete_talent_management(request, TalentID):
    talent_management = get_object_or_404(RiskTalent, TalentID=TalentID) 
    talent_management.delete()
    messages.success(request, "Risk Talent deleted successfully!")
    return redirect('list_talent_management')


#Business Context Analyzer
def list_business_context(request):
    # Get all records by default
    business_contexts = BusinessContext.objects.all()
    form = BusinessContextForm()

    
    # Default sorting by context_id
    sort_by = request.GET.get('sort_by', 'context_id')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        business_contexts = business_contexts.order_by(sort_by)
    else:
        business_contexts = business_contexts.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(business_contexts, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_business_contexts = paginator.page(page)
    except PageNotAnInteger:
        paginated_business_contexts = paginator.page(1)
    except EmptyPage:
        paginated_business_contexts = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Business Context Analyzer",
        'business_contexts':business_contexts, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_business_contexts': paginated_business_contexts,
        'rows_per_page': rows_per_page,
        'form': form,
    }
    return render(request, 'strategic_planning/business_context.html', context)


# Add Oversigh 

def add_business_context(request):
    form = BusinessContextForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Business Context added successfully!")
            return redirect('list_business_context')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Business Context. Please correct the errors.")
    return redirect('list_business_context')

# Business Context Dashboard
def edit_business_context(request, context_id):
    business_context = get_object_or_404(BusinessContext, context_id=context_id)
    if request.method == 'POST':
        form = BusinessContextForm(request.POST, instance=business_context)
        if form.is_valid():
            form.save()
            messages.success(request, "Business Context updated successfully!")
            return redirect('list_business_context')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Business Context. Please correct the errors.")
    else:
        form = BusinessContextForm(instance=business_context)
    return render(request, 'strategic_planning/business_context.html', {'form': form})


# Delete business_context
def delete_business_context(request, context_id):
    business_context = get_object_or_404(BusinessContext, context_id=context_id) 
    business_context.delete()
    messages.success(request, "Business Context deleted successfully!")
    return redirect('list_business_context')


# Risk Appetite
def list_risk_appetite(request):
    # Get all records by default
    risk_appetites = RiskAppetite.objects.all()
    form = RiskAppetiteForm()

    
    # Default sorting by appetite_id
    sort_by = request.GET.get('sort_by', 'appetite_id')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        risk_appetites = risk_appetites.order_by(sort_by)
    else:
        risk_appetites = risk_appetites.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(risk_appetites, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_risk_appetites = paginator.page(page)
    except PageNotAnInteger:
        paginated_risk_appetites = paginator.page(1)
    except EmptyPage:
        paginated_risk_appetites = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Risk Appetite",
        'risk_appetites':risk_appetites, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_risk_appetites': paginated_risk_appetites,
        'rows_per_page': rows_per_page,
        'form': form,
        'approved_bys': Staff.objects.all(),
        'review_cycles': RiskAppetite._meta.get_field('review_cycle').choices,
    }
    return render(request, 'strategic_planning/risk_appetite.html', context)


# Add Oversigh 

def add_risk_appetite(request):
    form = RiskAppetiteForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Appetite added successfully!")
            return redirect('list_risk_appetite')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Risk Appetite. Please correct the errors.")
    return redirect('list_risk_appetite')

# RiskAppetite Dashboard
def edit_risk_appetite(request, appetite_id):
    risk_appetite = get_object_or_404(RiskAppetite, appetite_id=appetite_id)
    if request.method == 'POST':
        form = RiskAppetiteForm(request.POST, instance=risk_appetite)
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Appetite updated successfully!")
            return redirect('list_risk_appetite')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Risk Appetite. Please correct the errors.")
    else:
        form = RiskAppetiteForm(instance=risk_appetite)
    return render(request, 'strategic_planning/risk_appetite.html', {'form': form})


# Delete risk_appetite
def delete_risk_appetite(request, appetite_id):
    risk_appetite = get_object_or_404(RiskAppetite, appetite_id=appetite_id) 
    risk_appetite.delete()
    messages.success(request, "Risk Appetite deleted successfully!")
    return redirect('list_risk_appetite')

# Strategic Evaluation
def list_strategic_evaluation(request):
    # Get all records by default
    strategic_evaluations = StrategicEvaluation.objects.all()
    form = StrategicEvaluationForm()

    
    # Default sorting by evaluation_id
    sort_by = request.GET.get('sort_by', 'evaluation_id')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        strategic_evaluations = strategic_evaluations.order_by(sort_by)
    else:
        strategic_evaluations = strategic_evaluations.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(strategic_evaluations, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_strategic_evaluations = paginator.page(page)
    except PageNotAnInteger:
        paginated_strategic_evaluations = paginator.page(1)
    except EmptyPage:
        paginated_strategic_evaluations = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Strategic Evaluation Tool",
        'strategic_evaluations':strategic_evaluations, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_strategic_evaluations': paginated_strategic_evaluations,
        'rows_per_page': rows_per_page,
        'form': form,
    }
    return render(request, 'strategic_planning/strategic_evaluation.html', context)


# Add Oversigh 

def add_strategic_evaluation(request):
    form = StrategicEvaluationForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Strategic Evaluation added successfully!")
            return redirect('list_strategic_evaluation')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Strategic Evaluation. Please correct the errors.")
    return redirect('list_strategic_evaluation')

# Strategic Evaluation Dashboard
def edit_strategic_evaluation(request, evaluation_id):
    strategic_evaluation = get_object_or_404(StrategicEvaluation, evaluation_id=evaluation_id)
    if request.method == 'POST':
        form = StrategicEvaluationForm(request.POST, instance=strategic_evaluation)
        if form.is_valid():
            form.save()
            messages.success(request, "Strategic Evaluation updated successfully!")
            return redirect('list_strategic_evaluation')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Strategic Evaluation. Please correct the errors.")
    else:
        form = StrategicEvaluationForm(instance=strategic_evaluation)
    return render(request, 'strategic_planning/strategic_evaluation.html', {'form': form})


# Delete strategic_evaluation
def delete_strategic_evaluation(request, evaluation_id):
    strategic_evaluation = get_object_or_404(StrategicEvaluation, evaluation_id=evaluation_id) 
    strategic_evaluation.delete()
    messages.success(request, "Strategic Evaluation deleted successfully!")
    return redirect('list_strategic_evaluation')

def list_objective(request):
    # Get all records by default
    objectives = Objective.objects.all()
    form = ObjectiveForm()

    
    # Default sorting by id
    sort_by = request.GET.get('sort_by', 'id')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        objectives = objectives.order_by(sort_by)
    else:
        objectives = objectives.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(objectives, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_objectives = paginator.page(page)
    except PageNotAnInteger:
        paginated_objectives = paginator.page(1)
    except EmptyPage:
        paginated_objectives = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Objectives",
        'objectives':objectives, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_objectives': paginated_objectives,
        'rows_per_page': rows_per_page,
        'form': form,
        'Status_choices': Objective._meta.get_field('progress_status').choices,
    }
    return render(request, 'strategic_planning/objective.html', context)


# Add Oversigh 

def add_objective(request):
    form = ObjectiveForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Objective added successfully!")
            return redirect('list_objective')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Objective. Please correct the errors.")
    return redirect('list_objective')

# Objective Dashboard
def edit_objective(request, id):
    objective = get_object_or_404(Objective, id=id)
    if request.method == 'POST':
        form = ObjectiveForm(request.POST, instance=objective)
        if form.is_valid():
            form.save()
            messages.success(request, "Objective updated successfully!")
            return redirect('list_objective')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Objective. Please correct the errors.")
    else:
        form = ObjectiveForm(instance=objective)
    return render(request, 'strategic_planning/objective.html', {'form': form})


# Delete objective
def delete_objective(request, id):
    objective = get_object_or_404(Objective, id=id) 
    objective.delete()
    messages.success(request, "Objective deleted successfully!")
    return redirect('list_objective')


# Continuous monitoring
# Change Assessment
def list_change_assessment(request):
    # Get all records by default
    change_assessments = ChangeAssessment.objects.all()
    form = ChangeAssessmentForm()

    
    # Default sorting by change_id
    sort_by = request.GET.get('sort_by', 'change_id')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        change_assessments = change_assessments.order_by(sort_by)
    else:
        change_assessments = change_assessments.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(change_assessments, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_change_assessments = paginator.page(page)
    except PageNotAnInteger:
        paginated_change_assessments = paginator.page(1)
    except EmptyPage:
        paginated_change_assessments = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Change Assessment",
        'change_assessments':change_assessments, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_change_assessments': paginated_change_assessments,
        'rows_per_page': rows_per_page,
        'form': form,
        'responsible_partys': Staff.objects.all(),
        'Status_choices': ChangeAssessment._meta.get_field('status').choices,
        'change_types': ChangeAssessment._meta.get_field('change_type').choices,
    }
    return render(request, 'continuous_monitoring/change_assessment.html', context)


# Add Oversigh 

def add_change_assessment(request):
    form = ChangeAssessmentForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Assessment added successfully!")
            return redirect('list_change_assessment')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Assessment. Please correct the errors.")
    return redirect('list_change_assessment')

# Change Assessment Dashboard
def edit_change_assessment(request, change_id):
    change_assessment = get_object_or_404(ChangeAssessment, change_id=change_id)
    if request.method == 'POST':
        form = ChangeAssessmentForm(request.POST, instance=change_assessment)
        if form.is_valid():
            form.save()
            messages.success(request, "Change Assessment updated successfully!")
            return redirect('list_change_assessment')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Change Assessment. Please correct the errors.")
    else:
        form = ChangeAssessmentForm(instance=change_assessment)
    return render(request, 'continuous_monitoring/change_assessment.html', {'form': form})



# Delete change_assessment
def delete_change_assessment(request, change_id):
    change_assessment = get_object_or_404(ChangeAssessment, change_id=change_id) 
    change_assessment.delete()
    messages.success(request, "Change Assessment deleted successfully!")
    return redirect('list_change_assessment')

def list_performance_review(request):
    # Get all records by default
    performance_reviews = PerformanceReview.objects.all()
    form = PerformanceReviewForm()

    
    # Default sorting by review_id
    sort_by = request.GET.get('sort_by', 'review_id')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        performance_reviews = performance_reviews.order_by(sort_by)
    else:
        performance_reviews = performance_reviews.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(performance_reviews, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_performance_reviews = paginator.page(page)
    except PageNotAnInteger:
        paginated_performance_reviews = paginator.page(1)
    except EmptyPage:
        paginated_performance_reviews = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Performance Review",
        'performance_reviews':performance_reviews, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_performance_reviews': paginated_performance_reviews,
        'rows_per_page': rows_per_page,
        'form': form,
        'reviewers': Staff.objects.all(),
        'objectives': Objective.objects.all(),
    }
    return render(request, 'continuous_monitoring/performance_review.html', context)


# Add Oversigh 

def add_performance_review(request):
    form = PerformanceReviewForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Assessment added successfully!")
            return redirect('list_performance_review')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Assessment. Please correct the errors.")
    return redirect('list_performance_review')

# Performance Review Dashboard
def edit_performance_review(request, review_id):
    performance_review = get_object_or_404(PerformanceReview, review_id=review_id)
    if request.method == 'POST':
        form = PerformanceReviewForm(request.POST, instance=performance_review)
        if form.is_valid():
            form.save()
            messages.success(request, "Performance Review updated successfully!")
            return redirect('list_performance_review')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Performance Review. Please correct the errors.")
    else:
        form = PerformanceReviewForm(instance=performance_review)
    return render(request, 'continuous_monitoring/performance_review.html', {'form': form})



# Delete performance_review
def delete_performance_review(request, review_id):
    performance_review = get_object_or_404(PerformanceReview, review_id=review_id) 
    performance_review.delete()
    messages.success(request, "Performance Review deleted successfully!")
    return redirect('list_performance_review')


#Improvement Action

def list_improvement_action(request):
    # Get all records by default
    improvement_actions = ImprovementAction.objects.all()
    form = ImprovementActionForm()

    
    # Default sorting by improvement_id
    sort_by = request.GET.get('sort_by', 'improvement_id')
    order = request.GET.get('order', 'asc')

    # Toggle order on each click
    if order == 'asc':
        improvement_actions = improvement_actions.order_by(sort_by)
    else:
        improvement_actions = improvement_actions.order_by(F(sort_by).desc())

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(improvement_actions, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_improvement_actions = paginator.page(page)
    except PageNotAnInteger:
        paginated_improvement_actions = paginator.page(1)
    except EmptyPage:
        paginated_improvement_actions = paginator.page(paginator.num_pages)

    
    # Pass necessary context to the template
    context = {
        'page_title': "Improvement Action",
        'improvement_actions':improvement_actions, 
        'current_sort': sort_by,
        'current_order': order,       
        'paginated_improvement_actions': paginated_improvement_actions,
        'rows_per_page': rows_per_page,
        'form': form,
        'initiated_bys': Staff.objects.all(),
        'related_risks': Risk.objects.all(),
        'Status_choices': ImprovementAction._meta.get_field('current_status').choices,
    }
    return render(request, 'continuous_monitoring/improvement_action.html', context)


# Add Oversigh 

def add_improvement_action(request):
    form = ImprovementActionForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Assessment added successfully!")
            return redirect('list_improvement_action')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Assessment. Please correct the errors.")
    return redirect('list_improvement_action')

# Improvement Action Dashboard
def edit_improvement_action(request, improvement_id):
    improvement_action = get_object_or_404(ImprovementAction, improvement_id=improvement_id)
    if request.method == 'POST':
        form = ImprovementActionForm(request.POST, instance=improvement_action)
        if form.is_valid():
            form.save()
            messages.success(request, "Improvement Action updated successfully!")
            return redirect('list_improvement_action')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Improvement Action. Please correct the errors.")
    else:
        form = ImprovementActionForm(instance=improvement_action)
    return render(request, 'continuous_monitoring/improvement_action.html', {'form': form})



# Delete improvement_action
def delete_improvement_action(request, improvement_id):
    improvement_action = get_object_or_404(ImprovementAction, improvement_id=improvement_id) 
    improvement_action.delete()
    messages.success(request, "Improvement Action deleted successfully!")
    return redirect('list_improvement_action')

#------------------------------
# Old Views
#------------------------------

""" # Add Objective View
def add_objective(request):
    form = ObjectiveForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Objective added successfully!")
            return redirect('list_objectives')
        else:
            messages.error(request, "Failed to add objective. Please correct the errors.")
    return render(request, 'business_objectives/list_objectives.html', {'form': form})

# Edit Objective View
def edit_objective(request, id):
    instance = get_object_or_404(Objective, id=id)
    form = ObjectiveForm(request.POST or None, instance=instance)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Objective updated successfully!")
            return redirect('list_objectives')
        else:
            form = ObjectiveForm(instance=instance)
            messages.error(request, "Failed to update objective. Please correct the errors.")
    return render(request, 'business_objectives/list_objectives.html', {'form': form, 'instance':instance,})


    
# Delete Objectives View
def delete_objective(request, id):
    objective = get_object_or_404(Objective, id=id)
    objective.delete()
    messages.success(request, "Objective deleted successfully!")
    return redirect('list_objectives')

# List Objectives View
PlaceholderObjective = namedtuple(
    'PlaceholderObjective',
    ['id', 'name', 'owner', 'department', 'start_date', 'end_date', 'status', 'priority_level']
)

def list_objectives(request):
    objectives = Objective.objects.all()  # Fetch all objectives from the database
    departments = Department.objects.all()
    owners = Staff.objects.all()
    form = ObjectiveForm()
    # Pagination setup
    # Get rows_per_page from the query parameters; default to 10
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10  # Fallback in case of invalid input

    # Set up pagination
    paginator = Paginator(objectives, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_objectives = paginator.page(page)
    except PageNotAnInteger:
        paginated_objectives = paginator.page(1)
    except EmptyPage:
        paginated_objectives = paginator.page(paginator.num_pages)

    # Add a blank placeholder objective to the paginated list
    blank_objective = PlaceholderObjective(
        id=1,
        name=None,
        owner=None,
        department=None,
        start_date=None,
        end_date=None,
        status=None,
        priority_level=None,
    )
    objectives_with_placeholder = [blank_objective] + list(paginated_objectives)

    context = {
        'page_title': "Objectives",
        'objectives': objectives_with_placeholder,
        'paginated_objectives': paginated_objectives,  # Pass the paginator object separately
        'rows_per_page': str(rows_per_page),
        'request': request,
        'status_choices': Objective.STATUS_CHOICES,
        'priority_choices': Objective.PRIORITY_CHOICES,
        'progress_status': Objective.PROGRESS_CHOICES,
        'form': form,
        'owners': owners,
        'departments': departments,  
    }
    return render(request, 'business_objectives/list_objectives.html', context)

# Strategy
def list_objective_strategies(request):
    strategies = ObjectiveStrategy.objects.all()  # Get all strategies
    objectives = Objective.objects.all()  # Get all objectives
    owners = Staff.objects.all()
    form = ObjectiveStrategyForm()

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(strategies, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_strategies = paginator.page(page)
    except PageNotAnInteger:
        paginated_strategies = paginator.page(1)
    except EmptyPage:
        paginated_strategies = paginator.page(paginator.num_pages)

    # Add a blank placeholder strategy
    PlaceholderStrategy = namedtuple(
        'PlaceholderStrategy',
        ['id', 'strategy_name', 'objective', 'owner', 'timeline', 'strategy_description']
    )
    blank_strategy = PlaceholderStrategy(
        id=1,
        strategy_name=None,
        objective=None,
        owner=None,
        timeline=None,
        strategy_description=None,
    )
    strategies_with_placeholder = [blank_strategy] + list(paginated_strategies)

    context = {
        'strategies': strategies_with_placeholder,
        'paginated_strategies': paginated_strategies,  # Pass the paginator object
        'rows_per_page': rows_per_page,
        'objectives': objectives,
        'owners': owners,
        'form': form,
    }
    return render(request, 'business_objectives/list_strategies.html', context)

def add_objective_strategy(request):
    form = ObjectiveStrategyForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Objective Strategy added successfully!")
            return redirect('list_objective_strategies')
        else:
            messages.error(request, "Failed to add Objective Strategy. Please correct the errors.")
    return redirect('list_objective_strategies')


def edit_objective_strategy(request, id):
    strategy = get_object_or_404(ObjectiveStrategy, id=id)  # Correctly fetch strategy by ID

    if request.method == 'POST':
        form = ObjectiveStrategyForm(request.POST, instance=strategy)
        if form.is_valid():
            form.save()
            messages.success(request, "Objective Strategy updated successfully!")
            return redirect('list_objective_strategies')
        else:
            messages.error(request, "Failed to update Objective Strategy. Please correct the errors.")
    else:
        form = ObjectiveStrategyForm(instance=strategy)

    strategies = ObjectiveStrategy.objects.all()
    return render(request, 'business_objectives/list_strategies.html', {
        'strategies': strategies,
        'form': form,
        'objectives': Objective.objects.all(),
        'owners': Staff.objects.all(),
    })


def delete_objective_strategy(request, id):
    strategy = get_object_or_404(ObjectiveStrategy, id=id)
    strategy.delete()
    messages.success(request, "Objective Strategy deleted successfully!")
    return redirect('list_objective_strategies')

#progress

# List all ObjectiveProgress entries

PlaceholderProgress = namedtuple(
    'PlaceholderProgress',
    ['id', 'objective', 'progress_percentage', 'last_updated', 'comments']
)

def list_objective_progress(request):
    progress_entries = ObjectiveProgress.objects.all()
    objectives = Objective.objects.all()  # Required for dropdown in Add Modal
    form = ObjectiveProgressForm()

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10  # Default if invalid input

    paginator = Paginator(progress_entries, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_progress = paginator.page(page)
    except PageNotAnInteger:
        paginated_progress = paginator.page(1)
    except EmptyPage:
        paginated_progress = paginator.page(paginator.num_pages)

    # Add a blank placeholder progress entry
    blank_progress = PlaceholderProgress(
        id=1,
        objective=None,
        progress_percentage=None,
        last_updated=None,
        comments=None
    )
    progress_with_placeholder = [blank_progress] + list(paginated_progress)

    context = {
        'page_title': "Objective Progress",
        'progress_entries': progress_with_placeholder,
        'paginated_progress': paginated_progress,  # Pass the paginator object
        'rows_per_page': rows_per_page,
        'form': form,  # Form for the Add Modal
        'objectives': objectives,  # Objectives for dropdown in Add Modal
    }
    return render(request, 'business_objectives/list_progress.html', context)


# Add a new ObjectiveProgress entry
def add_objective_progress(request):
    form = ObjectiveProgressForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Objective Progress added successfully!")
            return redirect('list_objective_progress')
        else:
            messages.error(request, "Failed to add Objective Progress. Please correct the errors.")
    return redirect('list_objective_progress')


# Edit an existing ObjectiveProgress entry
def edit_objective_progress(request, id):
    progress = get_object_or_404(ObjectiveProgress, id=id)
    form = ObjectiveProgressForm(request.POST or None, instance=progress)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Objective Progress updated successfully!")
            return redirect('list_objective_progress')
        else:
            messages.error(request, "Failed to update Objective Progress. Please correct the errors.")
    progress_entries = ObjectiveProgress.objects.all()
    return render(request, 'business_objectives/list_progress.html', {
        'progress_entries': progress_entries,
        'form': form,
        'objectives': Objective.objects.all(),
    })


# Delete an ObjectiveProgress entry
def delete_objective_progress(request, id):
    progress = get_object_or_404(ObjectiveProgress, id=id)
    progress.delete()
    messages.success(request, "Objective Progress deleted successfully!")
    return redirect('list_objective_progress')

#------------------------------
# Inherent Risk Track
#------------------------------

# List Inherent Risks

PlaceholderRisk = namedtuple(
    'PlaceholderRisk',
    [
        'risk_id',
        'name',
        'description',
        'category',
        'owner',
        'date_identified',
        'likelihood',
        'impact',
        'score',
        'risk_type',
        'source_of_risk',
        'associated_objective',
        'current_controls',
        'review_frequency',
    ]
)

def list_inherent_risks(request):
    inherent_risks = InherentRisk.objects.all()
    objectives = Objective.objects.all()
    owners = Staff.objects.all()
    form = InherentRiskForm()

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10

    paginator = Paginator(inherent_risks, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_risks = paginator.page(page)
    except PageNotAnInteger:
        paginated_risks = paginator.page(1)
    except EmptyPage:
        paginated_risks = paginator.page(paginator.num_pages)

    # Add a placeholder for a blank row
    blank_risk = PlaceholderRisk(
        risk_id=999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999,
        name=None,
        description=None,
        category="",
        owner=None,
        date_identified=None,
        likelihood="",
        impact="",
        score=None,
        risk_type="",
        source_of_risk="",
        associated_objective=None,
        current_controls="",
        review_frequency="",
    )
    risks_with_placeholder = [blank_risk] + list(paginated_risks)

    context = {
        'page_title': "Inherent Risks",
        'inherent_risks': risks_with_placeholder,
        'paginated_risks': paginated_risks,
        'rows_per_page': rows_per_page,
        'form': form,
        'objectives': objectives,
        'owners': owners,
        'review_frequency_choices': InherentRisk.REVIEW_FREQUENCY_CHOICES,
        'risk_type_choices': InherentRisk.RISK_TYPE_CHOICES,
        'impact_choices': InherentRisk.RISK_IMPACT_CHOICES,
        'likelihood_choices': InherentRisk.RISK_LIKELIHOOD_CHOICES,
    }
    return render(request, 'inherent_risk_track/list_inherent_risks.html', context)


# Add Inherent Risk
def add_inherent_risk(request):
    form = InherentRiskForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Inherent Risk added successfully!")
            return redirect('list_inherent_risks')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Inherent Risk. Please correct the errors.")
    return redirect('list_inherent_risks')

# Edit Inherent Risk
def edit_inherent_risk(request, id):
    inherent_risk = get_object_or_404(InherentRisk, risk_id=id)
    if request.method == 'POST':
        form = InherentRiskForm(request.POST, instance=inherent_risk)
        if form.is_valid():
            form.save()
            messages.success(request, "Inherent Risk updated successfully!")
            return redirect('list_inherent_risks')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Inherent Risk. Please correct the errors.")
    else:
        form = InherentRiskForm(instance=inherent_risk)
    return render(request, 'inherent_risk_track/list_inherent_risks.html', {'form': form})



# Delete Inherent Risk
def delete_inherent_risk(request, id):
    inherent_risk = get_object_or_404(InherentRisk, risk_id=id)  # Use `risk_id` instead of `id`
    inherent_risk.delete()
    messages.success(request, "Inherent Risk deleted successfully!")
    return redirect('list_inherent_risks')


# Risk Assessment

# List View
def list_risk_assessments(request):
    risk_assessments = RiskAssessment.objects.all()
    inherent_risks = InherentRisk.objects.all()
    assessed_bys = Staff.objects.all()
    form = RiskAssessmentForm()

    # Pagination setup
    paginator = Paginator(risk_assessments, 10)
    page = request.GET.get('page', 1)
    try:
        paginated_assessments = paginator.page(page)
    except PageNotAnInteger:
        paginated_assessments = paginator.page(1)
    except EmptyPage:
        paginated_assessments = paginator.page(paginator.num_pages)

    # Placeholder entry
    blank_assessment = RiskAssessment(
        assessment_id=1,
        risk=None,
        assessment_date=None,
        assessed_by=None,
        risk_value=None,
        probability_score=None,
        impact_score=None,
    )
    assessments_with_placeholder = [blank_assessment] + list(paginated_assessments)

    context = {
        'page_title': "Risk Assessments",
        'risk_assessments': assessments_with_placeholder,
        'paginated_assessments': paginated_assessments,
        'inherent_risks': inherent_risks,
        'assessed_bys': assessed_bys,
        'form': form,
    }
    return render(request, 'inherent_risk_track/risk_assessment.html', context)

# Add View
def add_risk_assessment(request):
    form = RiskAssessmentForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Assessment added successfully!")
            return redirect('list_risk_assessments')
        else:
            messages.error(request, "Failed to add Risk Assessment. Please correct the errors.")
    return redirect('list_risk_assessments')

# Edit View
def edit_risk_assessment(request, assessment_id):
    risk_assessment = get_object_or_404(RiskAssessment, assessment_id=assessment_id)
    if request.method == 'POST':
        form = RiskAssessmentForm(request.POST, instance=risk_assessment)
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Assessment updated successfully!")
            return redirect('list_risk_assessments')
        else:
            messages.error(request, "Failed to update Risk Assessment. Please correct the errors.")
    else:
        form = RiskAssessmentForm(instance=risk_assessment)
    return render(request, 'risk_assessment/list_risk_assessments.html', {
        'form': form,
        'risk_assessments': RiskAssessment.objects.all(),
    })


# Delete View
def delete_risk_assessment(request, assessment_id):
    risk_assessment = get_object_or_404(RiskAssessment, assessment_id=assessment_id)
    risk_assessment.delete()
    messages.success(request, "Risk Assessment deleted successfully!")
    return redirect('list_risk_assessments')


#------------------------------
# Key Risk Indicators (KRIs)
#------------------------------


# Add Key Risk Indicator View
def add_key_risk_indicator(request):
    form = KeyRiskIndicatorForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Key Risk Indicator added successfully!")
            return redirect('list_key_risk_indicators')
        else:
            messages.error(request, "Failed to add Key Risk Indicator. Please correct the errors.")
    return redirect('list_key_risk_indicators')


# Edit Key Risk Indicator View
def edit_key_risk_indicator(request, kri_id):
    key_risk_indicator = get_object_or_404(KeyRiskIndicator, kri_id=kri_id)
    if request.method == 'POST':
        form = KeyRiskIndicatorForm(request.POST, instance=key_risk_indicator)
        if form.is_valid():
            form.save()
            messages.success(request, "Key Risk Indicator updated successfully!")
            return redirect('list_key_risk_indicators')
        else:
            messages.error(request, "Failed to update Key Risk Indicator. Please correct the errors.")
    else:
        form = KeyRiskIndicatorForm(instance=key_risk_indicator)
    context = {
        'form': form,
        'key_risk_indicators': KeyRiskIndicator.objects.all(),
        'measurement_frequency_choices': KeyRiskIndicator._meta.get_field('measurement_frequency').choices,
        'status_choices': KeyRiskIndicator._meta.get_field('status').choices,
        'alert_triggers_choices': KeyRiskIndicator._meta.get_field('alert_triggers').choices,
        'inherent_risks': InherentRisk.objects.all(),
    }
    return render(request, 'key_risk_indicators/list_key_risk_indicators.html', context)


# Delete Key Risk Indicator View
def delete_key_risk_indicator(request, kri_id):
    key_risk_indicator = get_object_or_404(KeyRiskIndicator, kri_id=kri_id)
    key_risk_indicator.delete()
    messages.success(request, "Key Risk Indicator deleted successfully!")
    return redirect('list_key_risk_indicators')


# List Key Risk Indicators View
def list_key_risk_indicators(request):
    key_risk_indicators = KeyRiskIndicator.objects.all()
    inherent_risks = InherentRisk.objects.all()
    form = KeyRiskIndicatorForm()
    owners = Staff.objects.all()

    # Pagination setup
    paginator = Paginator(key_risk_indicators, 10)
    page = request.GET.get('page', 1)

    try:
        paginated_kris = paginator.page(page)
    except PageNotAnInteger:
        paginated_kris = paginator.page(1)
    except EmptyPage:
        paginated_kris = paginator.page(paginator.num_pages)

    # Placeholder entry
    placeholder_kri = KeyRiskIndicator(
        kri_id=1,
        name=None,
        related_risk=None,
        current_value=None,
        threshold_lower=None,
        threshold_upper=None,
        measurement_frequency=None,
        status=None,
        owner=None,
        data_source=None,
        description=None,
        alert_triggers=None,
    )

    kris_with_placeholder = [placeholder_kri] + list(paginated_kris)

    context = {
        'page_title': "Key Risk Indicators",
        'paginated_kris': paginated_kris,
        'kris_with_placeholder': kris_with_placeholder,
        'form': form,
        'owners': owners,
        'measurement_frequency_choices': KeyRiskIndicator._meta.get_field('measurement_frequency').choices,
        'status_choices': KeyRiskIndicator._meta.get_field('status').choices,
        'alert_triggers_choices': KeyRiskIndicator._meta.get_field('alert_triggers').choices,
        'inherent_risks': inherent_risks,
    }
    return render(request, 'key_risk_indicators/list_key_risk_indicators.html', context)


# Key Metrics

def list_key_metrics(request):
    key_metrics = KeyMetric.objects.all()
    key_risk_indicators = KeyRiskIndicator.objects.all()
    form = KeyMetricForm()

    # Pagination setup
    paginator = Paginator(key_metrics, 10)  # Show 10 metrics per page
    page = request.GET.get('page', 1)
    try:
        paginated_metrics = paginator.page(page)
    except PageNotAnInteger:
        paginated_metrics = paginator.page(1)
    except EmptyPage:
        paginated_metrics = paginator.page(paginator.num_pages)

    # Placeholder row
    blank_metric = KeyMetric(
        metric_id=1,
        name=None,
        description=None,
        related_kri=None,
        date_recorded=None,
        comment=None,
    )
    metrics_with_placeholder = [blank_metric] + list(paginated_metrics)

    context = {
        'page_title': "Key Metrics",
        'key_metrics': metrics_with_placeholder,
        'paginated_metrics': paginated_metrics,
        'key_risk_indicators': key_risk_indicators,
        'form': form,
    }
    return render(request, 'key_risk_indicators/list_key_metrics.html', context)

def add_key_metric(request):
    form = KeyMetricForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Key Metric added successfully!")
            return redirect('list_key_metrics')
        else:
            messages.error(request, "Failed to add Key Metric. Please correct the errors.")
    return redirect('list_key_metrics')

def edit_key_metric(request, metric_id):
    key_metric = get_object_or_404(KeyMetric, metric_id=metric_id)
    if request.method == 'POST':
        form = KeyMetricForm(request.POST, instance=key_metric)
        if form.is_valid():
            form.save()
            messages.success(request, "Key Metric updated successfully!")
            return redirect('list_key_metrics')
        else:
            messages.error(request, "Failed to update Key Metric. Please correct the errors.")
    else:
        form = KeyMetricForm(instance=key_metric)
    context = {
        'form': form,
        'key_metrics': KeyMetric.objects.all(),
        'key_risk_indicators': KeyRiskIndicator.objects.all(),
    }
    return render(request, 'key_risk_indicators/list_key_metrics.html', context)

def delete_key_metric(request, metric_id):
    key_metric = get_object_or_404(KeyMetric, metric_id=metric_id)
    key_metric.delete()
    messages.success(request, "Key Metric deleted successfully!")
    return redirect('list_key_metrics')

#Risk Appetite Console
def add_risk_appetite(request):
    form = RiskAppetiteForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Appetite added successfully!")
            return redirect('list_risk_appetites')
        else:
            messages.error(request, "Failed to add Risk Appetite. Please correct the errors.")
    return redirect('list_risk_appetites')

def edit_risk_appetite(request, appetite_id):
    risk_appetite = get_object_or_404(RiskAppetite, appetite_id=appetite_id)
    if request.method == 'POST':
        form = RiskAppetiteForm(request.POST, instance=risk_appetite)
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Appetite updated successfully!")
            return redirect('list_risk_appetites')
        else:
            messages.error(request, "Failed to update Risk Appetite. Please correct the errors.")
    else:
        form = RiskAppetiteForm(instance=risk_appetite)
    context = {
        'form': form,
        'risk_appetites': RiskAppetite.objects.all(),
    }
    return render(request, 'risk_appetite_console/list_risk_appetites.html', context)

def delete_risk_appetite(request, appetite_id):
    risk_appetite = get_object_or_404(RiskAppetite, appetite_id=appetite_id)
    risk_appetite.delete()
    messages.success(request, "Risk Appetite deleted successfully!")
    return redirect('list_risk_appetites')

def list_risk_appetites(request):
    risk_appetites = RiskAppetite.objects.all()
    form = RiskAppetiteForm()
    owners = Staff.objects.all()

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10  # Fallback to default

    paginator = Paginator(risk_appetites, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_appetites = paginator.page(page)
    except PageNotAnInteger:
        paginated_appetites = paginator.page(1)
    except EmptyPage:
        paginated_appetites = paginator.page(paginator.num_pages)

    # Placeholder entry
    placeholder_appetite = RiskAppetite(
        appetite_id=1,
        name=None,
        category=None,
        owner=None,
        approving_authority=None,
        min_threshold=None,
        max_threshold=None,
        review_date=None,
    )
    appetites_with_placeholder = [placeholder_appetite] + list(paginated_appetites)

    context = {
        'page_title': "Risk Appetites",
        'paginated_appetites': paginated_appetites,
        'appetites_with_placeholder': appetites_with_placeholder,
        'form': form,
        'owners': owners,
    }
    return render(request, 'risk_appetite_console/list_risk_appetites.html', context)

def add_risk_tolerance(request):
    form = RiskToleranceForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Tolerance added successfully!")
            return redirect('list_risk_tolerances')
        else:
            messages.error(request, "Failed to add Risk Tolerance. Please correct the errors.")
    return redirect('list_risk_tolerances')

def edit_risk_tolerance(request, tolerance_id):
    risk_tolerance = get_object_or_404(RiskTolerance, tolerance_id=tolerance_id)
    if request.method == 'POST':
        form = RiskToleranceForm(request.POST, instance=risk_tolerance)
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Tolerance updated successfully!")
            return redirect('list_risk_tolerances')
        else:
            messages.error(request, "Failed to update Risk Tolerance. Please correct the errors.")
    else:
        form = RiskToleranceForm(instance=risk_tolerance)
    context = {
        'form': form,
        'risk_tolerances': RiskTolerance.objects.all(),
    }
    return render(request, 'risk_appetite_console/list_risk_tolerances.html', context)

def delete_risk_tolerance(request, tolerance_id):
    risk_tolerance = get_object_or_404(RiskTolerance, tolerance_id=tolerance_id)
    risk_tolerance.delete()
    messages.success(request, "Risk Tolerance deleted successfully!")
    return redirect('list_risk_tolerances')



def list_risk_tolerances(request):
    # Fetch all risk tolerances
    risk_tolerances = RiskTolerance.objects.all()
    inherent_risks = InherentRisk.objects.all()
    risk_appetites = RiskAppetite.objects.all()
    form = RiskToleranceForm()

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page',  10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page =  10  # Default value in case of invalid input

    paginator = Paginator(risk_tolerances, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_tolerances = paginator.page(page)
    except PageNotAnInteger:
        paginated_tolerances = paginator.page(1)
    except EmptyPage:
        paginated_tolerances = paginator.page(paginator.num_pages)

    # Add a blank row placeholder
    blank_tolerance = RiskTolerance(
        tolerance_id=0,
        name=None,
        related_appetite=None,
        related_risk=None,
        accepted_tolerance_level=None,
        monitoring_mechanism=None,
        current_status=None,
        review_date=None,
    )
    tolerances_with_placeholder = [blank_tolerance] + list(paginated_tolerances)

    # Context for template
    context = {
        'page_title': "Risk Tolerances",
        'risk_tolerances': tolerances_with_placeholder,
        'paginated_tolerances': paginated_tolerances,
        'rows_per_page': rows_per_page,
        'inherent_risks': inherent_risks,
        'risk_appetites': risk_appetites,
        'form': form,
        'current_status_choices': RiskTolerance._meta.get_field('current_status').choices,
    }
    return render(request, 'risk_appetite_console/list_risk_tolerances.html', context)


#Control Strength
# Add Control View
def add_control(request):
    form = ControlForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Control added successfully!")
            return redirect('list_controls')
        else:
            messages.error(request, "Failed to add control. Please correct the errors.")
    return redirect('list_controls')

# Edit Control View
def edit_control(request, control_id):
    control = get_object_or_404(Control, control_id=control_id)
    if request.method == 'POST':
        form = ControlForm(request.POST, instance=control)
        if form.is_valid():
            form.save()
            messages.success(request, "Control updated successfully!")
            return redirect('list_controls')
        else:
            messages.error(request, "Failed to update control. Please correct the errors.")
    else:
        form = ControlForm(instance=control)
    context = {
        'form': form,
        'controls': Control.objects.all(),
    }
    return render(request, 'controls_effectiveness/list_controls.html', context)

# Delete Control View
def delete_control(request, control_id):
    control = get_object_or_404(Control, control_id=control_id)
    control.delete()
    messages.success(request, "Control deleted successfully!")
    return redirect('list_controls')

# List Controls View

def list_controls(request):
    controls = Control.objects.all()
    form = ControlForm()
    owners = Staff.objects.all()

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page',  10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page =  10  # Default to  10 if invalid input

    paginator = Paginator(controls, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_controls = paginator.page(page)
    except PageNotAnInteger:
        paginated_controls = paginator.page(1)
    except EmptyPage:
        paginated_controls = paginator.page(paginator.num_pages)

    # Add placeholder row
    blank_control = Control(
        control_id=0,
        name=None,
        related_risk=None,
        control_type=None,
        implementation_date=None,
        review_date=None,
        owner=None,
        control_strength=None,
    )
    controls_with_placeholder = [blank_control] + list(paginated_controls)

    context = {
        'page_title': "Controls",
        'controls': controls_with_placeholder,
        'paginated_controls': paginated_controls,
        'rows_per_page': rows_per_page,
        'form': form,
        'owners': owners,
        'inherent_risks': InherentRisk.objects.all(),
        'control_type_choices': Control._meta.get_field('control_type').choices,
        'control_strength_choices': Control._meta.get_field('control_strength').choices,
    }
    return render(request, 'controls_effectiveness/list_controls.html', context)


# Add Control Assessment View
def add_control_assessment(request):
    form = ControlAssessmentForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Control Assessment added successfully!")
            return redirect('list_control_assessments')
        else:
            messages.error(request, "Failed to add Control Assessment. Please correct the errors.")
    return redirect('list_control_assessments')

# Edit Control Assessment View
def edit_control_assessment(request, assessment_id):
    control_assessment = get_object_or_404(ControlAssessment, assessment_id=assessment_id)
    if request.method == 'POST':
        form = ControlAssessmentForm(request.POST, instance=control_assessment)
        if form.is_valid():
            form.save()
            messages.success(request, "Control Assessment updated successfully!")
            return redirect('list_control_assessments')
        else:
            messages.error(request, "Failed to update Control Assessment. Please correct the errors.")
    else:
        form = ControlAssessmentForm(instance=control_assessment)
    context = {
        'form': form,
        'control_assessments': ControlAssessment.objects.all(),
    }
    return render(request, 'controls_effectiveness/list_control_assessments.html', context)

# Delete Control Assessment View
def delete_control_assessment(request, assessment_id):
    control_assessment = get_object_or_404(ControlAssessment, assessment_id=assessment_id)
    control_assessment.delete()
    messages.success(request, "Control Assessment deleted successfully!")
    return redirect('list_control_assessments')

# Control Assessments
# List Control Assessments View

def list_control_assessments(request):
    control_assessments = ControlAssessment.objects.all()
    assessed_bys = Staff.objects.all()
    form = ControlAssessmentForm()

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page',  10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page =  10  # Default to  10 if invalid input

    paginator = Paginator(control_assessments, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_assessments = paginator.page(page)
    except PageNotAnInteger:
        paginated_assessments = paginator.page(1)
    except EmptyPage:
        paginated_assessments = paginator.page(paginator.num_pages)

    # Add placeholder row
    blank_assessment = ControlAssessment(
        assessment_id=0,
        name=None,
        related_control=None,
        effectiveness_rating=None,
        assessed_by=None,
        assessment_date=None,
        comments=None,
    )
    assessments_with_placeholder = [blank_assessment] + list(paginated_assessments)

    context = {
        'page_title': "Control Assessments",
        'control_assessments': assessments_with_placeholder,
        'paginated_assessments': paginated_assessments,
        'rows_per_page': rows_per_page,
        'form': form,
        'assessed_bys': assessed_bys,
        'controls': Control.objects.all(),
        'effectiveness_rating_choices': ControlAssessment._meta.get_field('effectiveness_rating').choices,
    }
    return render(request, 'controls_effectiveness/list_control_assessments.html', context)

#log
# Add Control Log View
def add_control_log(request):
    form = ControlLogForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Control Log added successfully!")
            return redirect('list_control_logs')
        else:
            messages.error(request, "Failed to add Control Log. Please correct the errors.")
    return redirect('list_control_logs')

# Edit Control Log View
def edit_control_log(request, log_id):
    control_log = get_object_or_404(ControlLog, log_id=log_id)
    form = ControlLogForm(request.POST or None, instance=control_log)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Control Log updated successfully!")
            return redirect('list_control_logs')
        else:
            messages.error(request, "Failed to update Control Log. Please correct the errors.")
    return render(request, 'controls_effectiveness/list_control_logs.html', {'form': form, 'control_logs': ControlLog.objects.all()})

# Delete Control Log View
def delete_control_log(request, log_id):
    control_log = get_object_or_404(ControlLog, log_id=log_id)
    control_log.delete()
    messages.success(request, "Control Log deleted successfully!")
    return redirect('list_control_logs')

# List Control Logs View

def list_control_logs(request):
    # Get all control logs
    control_logs = ControlLog.objects.all()
    controls = Control.objects.all()
    performed_bys = Staff.objects.all()

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10  # Default fallback

    paginator = Paginator(control_logs, rows_per_page)
    page = request.GET.get('page', 1)
    try:
        paginated_logs = paginator.page(page)
    except PageNotAnInteger:
        paginated_logs = paginator.page(1)
    except EmptyPage:
        paginated_logs = paginator.page(paginator.num_pages)

    # Add placeholder row for blank entry
    blank_log = ControlLog(
        log_id=0,
        name=None,
        activity=None,
        related_control=None,
        performed_by=None,
        timestamp=None,
    )
    logs_with_placeholder = [blank_log] + list(paginated_logs)

    # Initialize form for adding new logs
    form = ControlLogForm()

    context = {
        'page_title': "Control Logs",
        'control_logs': logs_with_placeholder,
        'paginated_logs': paginated_logs,
        'rows_per_page': rows_per_page,
        'form': form,
        'controls': controls,
        'performed_bys': performed_bys,
    }

    return render(request, 'controls_effectiveness/list_control_logs.html', context)

# Residual Risk Zone

# Add Residual Risk View
def add_residual_risk(request):
    form = ResidualRiskForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Residual Risk added successfully!")
            return redirect('list_residual_risks')
        else:
            messages.error(request, "Failed to add Residual Risk. Please correct the errors.")
    return redirect('list_residual_risks')

# Edit Residual Risk View
def edit_residual_risk(request, residual_risk_id):
    residual_risk = get_object_or_404(ResidualRisk, residual_risk_id=residual_risk_id)
    form = ResidualRiskForm(request.POST or None, instance=residual_risk)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Residual Risk updated successfully!")
            return redirect('list_residual_risks')
        else:
            print(form.errors)
            messages.error(request, "Failed to update Residual Risk. Please correct the errors.")
    
    context = {
        'form': form,
        'residual_risks': ResidualRisk.objects.all(),
    }
    return render(request, 'residual_risk_zone/list_residual_risks.html', context)

# Delete Residual Risk View
def delete_residual_risk(request, residual_risk_id):
    residual_risk = get_object_or_404(ResidualRisk, residual_risk_id=residual_risk_id)
    residual_risk.delete()
    messages.success(request, "Residual Risk deleted successfully!")
    return redirect('list_residual_risks')


# List Residual Risks View
def list_residual_risks(request):
    residual_risks = ResidualRisk.objects.all()
    owners = Staff.objects.all()
    
    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10  # Default rows per page

    paginator = Paginator(residual_risks, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_risks = paginator.page(page)
    except PageNotAnInteger:
        paginated_risks = paginator.page(1)
    except EmptyPage:
        paginated_risks = paginator.page(paginator.num_pages)

    # Placeholder entry
    blank_risk = ResidualRisk(
        residual_risk_id=0,
        related_inherent_risk=None,
        mitigation_actions=None,
        current_score=None,
        last_review_date=None,
        next_review_date=None,
        responsible_party=None,
        owner=None,
        residual_risk_rating=None,
    )
    risks_with_placeholder = [blank_risk] + list(paginated_risks)
    form = ResidualRiskForm()
    context = {
        'page_title': "Residual Risks",
        'residual_risks': risks_with_placeholder,
        'paginated_risks': paginated_risks,
        'rows_per_page': rows_per_page,
        'form': form,
        'owners': owners,
        'inherent_risks': InherentRisk.objects.all(),
        'risk_rating_choices': ResidualRisk._meta.get_field('residual_risk_rating').choices,
    }
    return render(request, 'residual_risk_zone/list_residual_risks.html', context)


# Add Residual Risk Assessment View
def add_residual_risk_assessment(request):
    form = ResidualRiskAssessmentForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Residual Risk Assessment added successfully!")
            return redirect('list_residual_risk_assessments')
        else:
            print(form.errors)
            messages.error(request, "Failed to add Residual Risk Assessment. Please correct the errors.")
    return redirect('list_residual_risk_assessments')

# Edit Residual Risk Assessment View
def edit_residual_risk_assessment(request, assessment_id):
    assessment = get_object_or_404(ResidualRiskAssessment, assessment_id=assessment_id)
    if request.method == 'POST':
        form = ResidualRiskAssessmentForm(request.POST, instance=assessment)
        if form.is_valid():
            form.save()
            messages.success(request, "Residual Risk Assessment updated successfully!")
            return redirect('list_residual_risk_assessments')
        else:
            messages.error(request, "Failed to update Residual Risk Assessment. Please correct the errors.")
    else:
        form = ResidualRiskAssessmentForm(instance=assessment)
    context = {
        'form': form,
        'residual_risk_assessments': ResidualRiskAssessment.objects.all(),
    }
    return render(request, 'residual_risk_zone/list_residual_risk_assessments.html', context)

# Delete Residual Risk Assessment View
def delete_residual_risk_assessment(request, assessment_id):
    assessment = get_object_or_404(ResidualRiskAssessment, assessment_id=assessment_id)
    assessment.delete()
    messages.success(request, "Residual Risk Assessment deleted successfully!")
    return redirect('list_residual_risk_assessments')

# List Residual Risk Assessments

def list_residual_risk_assessments(request):
    assessments = ResidualRiskAssessment.objects.all()
    assessed_bys = Staff.objects.all()
    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10  # Default rows per page

    paginator = Paginator(assessments, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_assessments = paginator.page(page)
    except PageNotAnInteger:
        paginated_assessments = paginator.page(1)
    except EmptyPage:
        paginated_assessments = paginator.page(paginator.num_pages)

    # Add a placeholder row at the beginning
    placeholder_assessment = ResidualRiskAssessment(
        assessment_id=0,
        related_residual_risk=None,
        updated_risk_value=None,
        updated_risk_rating=None,
        assessed_by=None,
        assessment_date=None,
    )
    assessments_with_placeholder = [placeholder_assessment] + list(paginated_assessments)

    form = ResidualRiskAssessmentForm()
    context = {
        'page_title': "Residual Risk Assessments",
        'residual_risk_assessments': assessments_with_placeholder,
        'paginated_assessments': paginated_assessments,
        'assessed_bys': assessed_bys,
        'form': form,
        'residual_risks': ResidualRisk.objects.all(),
        'risk_rating_choices': ResidualRiskAssessment._meta.get_field('updated_risk_rating').choices,
    }
    return render(request, 'residual_risk_zone/list_residual_risk_assessments.html', context)
# Remediation Actions

# Add Remediation Action View
def add_remediation_action(request):
    form = RemediationActionForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Remediation Action added successfully!")
            return redirect('list_remediation_actions')
        else:
            messages.error(request, "Failed to add Remediation Action. Please correct the errors.")
    return redirect('list_remediation_actions')


# Edit Remediation Action View
def edit_remediation_action(request, action_id):
    remediation_action = get_object_or_404(RemediationAction, action_id=action_id)
    if request.method == 'POST':
        form = RemediationActionForm(request.POST, instance=remediation_action)
        if form.is_valid():
            form.save()
            messages.success(request, "Remediation Action updated successfully!")
            return redirect('list_remediation_actions')
        else:
            messages.error(request, "Failed to update Remediation Action. Please correct the errors.")
    else:
        form = RemediationActionForm(instance=remediation_action)
    context = {
        'form': form,
        'remediation_actions': RemediationAction.objects.all(),
    }
    return render(request, 'remediation_actions/list_remediation_actions.html', context)


# Delete Remediation Action View
def delete_remediation_action(request, action_id):
    remediation_action = get_object_or_404(RemediationAction, action_id=action_id)
    remediation_action.delete()
    messages.success(request, "Remediation Action deleted successfully!")
    return redirect('list_remediation_actions')


# List Remediation Actions View

def list_remediation_actions(request):
    remediation_actions = RemediationAction.objects.all()
    owners = Staff.objects.all()

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10  # Default rows per page

    paginator = Paginator(remediation_actions, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_actions = paginator.page(page)
    except PageNotAnInteger:
        paginated_actions = paginator.page(1)
    except EmptyPage:
        paginated_actions = paginator.page(paginator.num_pages)

    # Add placeholder row
    blank_action = RemediationAction(
        action_id=0,
        related_risk=None,
        action_description=None,
        owner=None,
        start_date=None,
        due_date=None,
        status=None,
        progress_percentage=0,
        completion_date=None,
        associated_costs=0.0,
        comments=None,
    )
    actions_with_placeholder = [blank_action] + list(paginated_actions)

    form = RemediationActionForm()
    context = {
        'page_title': "Remediation Actions",
        'remediation_actions': actions_with_placeholder,
        'paginated_actions': paginated_actions,
        'rows_per_page': rows_per_page,
        'form': form,
        'owners': owners,
        'inherent_risks': InherentRisk.objects.all(),
        'status_choices': RemediationAction._meta.get_field('status').choices,
    }
    return render(request, 'remediation_actions/list_remediation_actions.html', context)


#Risk Radar

# Add Risk Radar
def add_risk_radar(request):
    form = RiskRadarForm(request.POST or None)
    if request.method == 'POST': 
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Radar entry added successfully!")
            return redirect('list_risk_radar')
        else:
            messages.error(request, "Failed to add Remediation Action. Please correct the errors.")
    return redirect('list_risk_radar')
    

# Edit Risk Radar
def edit_risk_radar(request, radar_id):
    radar = get_object_or_404(RiskRadar, radar_id=radar_id)
    if request.method == 'POST':
        form = RiskRadarForm(request.POST, instance=radar)
        if form.is_valid():
            form.save()
            messages.success(request, "Risk Radar entry updated successfully!")
            return redirect('list_risk_radar')
        else:
            print(form.errors)  # Debugging
            messages.error(request, f"Failed to update Risk Radar. Errors: {form.errors}")
    else:
        form = RiskRadarForm(instance=radar)

    context = {
        'form': form,
        'radar': radar,
        'risk_radars': RiskRadar.objects.all(),
        'inherent_risks': InherentRisk.objects.all(),
        'remediation_actions': RemediationAction.objects.all(),
        'zone_choices': RiskRadar._meta.get_field('zone').choices,
        'color_choices': RiskRadar._meta.get_field('color').choices,
    }
    return render(request, 'risk_radar/list_risk_radar.html', context)


# Delete Risk Radar
def delete_risk_radar(request, radar_id):
    radar = get_object_or_404(RiskRadar, radar_id=radar_id)
    radar.delete()
    messages.success(request, "Risk Radar entry deleted successfully!")
    return redirect('list_risk_radar')

# List Risk Radar
def list_risk_radar(request):
    # Fetch all risk radar entries
    radars = RiskRadar.objects.all()
    owners = Staff.objects.all()

    # Pagination setup
    rows_per_page = request.GET.get('rows_per_page', 10)
    try:
        rows_per_page = int(rows_per_page)
    except ValueError:
        rows_per_page = 10  # Default value

    paginator = Paginator(radars, rows_per_page)
    page = request.GET.get('page', 1)

    try:
        paginated_radars = paginator.page(page)
    except PageNotAnInteger:
        paginated_radars = paginator.page(1)
    except EmptyPage:
        paginated_radars = paginator.page(paginator.num_pages)

    # Placeholder row
    blank_radar = RiskRadar(
        radar_id=0,
        related_risk=None,
        related_action=None,
        description=None,
        owner=None,
        last_updated=None,
        zone=None,
        color=None,
    )
    radars_with_placeholder = [blank_radar] + list(paginated_radars)

    # Form for adding/editing radar entries
    form = RiskRadarForm()

    context = {
        'page_title': "Risk Radar",
        'risk_radars': radars_with_placeholder,
        'paginated_radars': paginated_radars,
        'form': form,
        'owners': owners,
        'inherent_risks': InherentRisk.objects.all(),
        'remediation_actions': RemediationAction.objects.all(),
        'zone_choices': RiskRadar._meta.get_field('zone').choices,
        'color_choices': RiskRadar._meta.get_field('color').choices,
        'rows_per_page': rows_per_page,
    }

    return render(request, 'risk_radar/list_risk_radar.html', context)
 """

#################################
#################################